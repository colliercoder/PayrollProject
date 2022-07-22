import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import holidays
from ColombianHolidays import days_before_holidays
from datetime import date
import datetime
import shutil
import os
DESKTOP = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

#Colombian Holidays
current_year = date.today().year
colombianHolidays = holidays.Colombia(years = current_year)
days_before_holidays = days_before_holidays()

# variabls for turno
turnoA = "A"
turnoB = "B"
turnoC = "C"

# variables for novedad
domingo = "DOMINGO"
festivo = "FESTIVO"

# variables for turno en que se efectua
domingo_dia = "DOMINGO DIA"
festivo_dia = "FESTIVO DIA"
domingo_noche = "DOMINGO NOCHE"
festivo_noche = "FESTIVO NOCHE"
sabado_noche = "SABADO NOCHE"
noche_antes_festivo = "NOCHE ANTES FESTIVO"

# variables for hora_en_que
ten_to_six = "10:00PM - 06:00AM"
ten_to_twelve = "10:00PM - 12:00AM"
twelve_to_six = "12:00AM-06:00AM"

# variables for tipo_de_recargo
nocturno = "NOCTURNO"
nocturno_dom_fest = "NOCTURNO DOMINICAL O FESTIVO"

class ExcelData:
    """INPUTING THE SCHEDULE FILE, THE CHOSEN SHEET, AND THE MONTH"""
    def __init__(self,schedule_file,sheet,month):
        global DESKTOP
        self.schedule_file = schedule_file #CHOSEN FILE (THE PATH)
        self.sheet = sheet
        self.month = month
        self.sheet = sheet
        original = r"files/hours_report_normal.xlsx"
        target = fr"{DESKTOP}\{self.month}_nominas_reporte.xlsx"
        shutil.copyfile(original,target)
        self.hours_report = xw.Book(target) #PATH OF HOURS REPORT

        """Establish a connection to a workbook, THE SCHEDULE FILE"""
        self.schedule_workbook = xw.Book(rf"{self.schedule_file}")
        self.wb = openpyxl.load_workbook(rf"{self.schedule_file}")

        """Instantiate the sheets objects"""
        self.current_schedule = self.schedule_workbook.sheets[self.sheet]
        self.attendance = self.hours_report.sheets['ATTENDANCE']
        self.domingo_festivo = self.hours_report.sheets['DOMINGO Y FESTIVO']
        self.recargo_nocturno = self.hours_report.sheets['RECARGO NOCTURNO']

        self.attendance.range("B2").value = f"REPORTE MENSUAL DE ATENDENCIA DE {month.upper()} DEL {current_year}"
        self.domingo_festivo.range("B2").value = f"REPORTE MENSUAL DE DOMINGOS Y FESTIVOS DE {month.upper()} DEL {current_year}"
        self.recargo_nocturno.range("B2").value = f"REPORTE MENSUAL DE RECARGO NOCTURNO DE {month.upper()} DEL {current_year}"

        self.sheet_picked = self.wb[self.sheet]
        self.lista = self.hours_report.sheets['LISTA']

        self.max_col = get_column_letter(self.sheet_picked.max_column)

    """------------------------------------------ATTENDANCE---------------------------------------------------------"""
    def attendance_creator(self,row_start = 9,col_start = 'E',name_col = 'D',cedula_col = 'C'):
        #initializing count and dictionary
        count = 0
        attendance_dict = {}
        for i in range(row_start - 1, self.sheet_picked.max_row + 1):  # looping down names
            name = name_col + str(i)
            cedula = cedula_col + str(i)

            name = self.current_schedule.range(name).value
            cedula = self.current_schedule.range(cedula).value

            for x in range(column_index_from_string(col_start), column_index_from_string(self.max_col) + 1, 1):  # looping through shifts
                shift = self.sheet_picked.cell(row=i, column=x).value
                the_date = self.sheet_picked.cell(row=row_start - 2, column=x).value
                if shift != 'O' and shift != 'D' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                    dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift}
                    count += 1
                    attendance_dict[count] = dicts
        print(attendance_dict)
        return attendance_dict

    def attendance_report(self,documento = 'C',nombre = 'D', shift = 'G', date = 'H',starting_row = 5):
        attendance_dict = self.attendance_creator()
        columnC = documento
        columnD = nombre
        columnG = shift
        columnH = date

        row = starting_row #starting row

        for entry in range(len(attendance_dict)):
            cellC = columnC + str(entry+row)
            cellD = columnD + str(entry+row)
            cellG = columnG + str(entry+row)
            cellH = columnH + str(entry+row)


            self.attendance.range(cellC).value = attendance_dict[entry+1]['cedula']
            self.attendance.range(cellD).value = attendance_dict[entry+1]['name']
            self.attendance.range(cellG).value = attendance_dict[entry+1]['shift']
            self.attendance.range(cellH).value = attendance_dict[entry+1]['date']

    """----------------------------------------DOMINGO FESTIVO-----------------------------------------------------"""

    def domingo_dict(self,row_start=9, col_start='E', name_col='D', cedula_col='C'):
        # initializing count and dictionary
        count = 0
        domingodict = {}

        for i in range(row_start - 1, self.sheet_picked.max_row + 1):  # looping down names
            name = name_col + str(i)
            cedula = cedula_col + str(i)

            name = self.current_schedule.range(name).value
            cedula = self.current_schedule.range(cedula).value

            for x in range(column_index_from_string(col_start), column_index_from_string(self.max_col) + 1,
                           1):  # looping through shifts
                shift = self.sheet_picked.cell(row=i, column=x).value
                the_date = self.sheet_picked.cell(row=row_start - 2, column=x).value
                if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                    if shift == 'C':
                        if the_date in colombianHolidays:  # Nightshift for a holiday
                            if ((the_date + datetime.timedelta(
                                    days=1)) in colombianHolidays):  # Holiday night shift with the next day being a holiday
                                dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                         'novedad': 'Festivo y Festivo manana', 'turno_en_que': festivo_noche,
                                         'num_of_horas': 8}
                                count += 1
                                domingodict[count] = dicts
                            elif the_date.weekday() + 1 == 6:  # Holiday night shift with sunday being the next day
                                dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                         'novedad': 'Festivo y Domingo manana', 'turno_en_que': festivo_noche,
                                         'num_of_horas': 8}
                                count += 1
                                domingodict[count] = dicts

                            else:
                                dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                         'novedad': festivo, 'turno_en_que': festivo_noche,
                                         'num_of_horas': 2}
                                count += 1
                                domingodict[count] = dicts

                        elif the_date.weekday() == 6:  # Sunday night
                            if ((the_date + datetime.timedelta(
                                    days=1)) in colombianHolidays):  # Sunday night with a holiday after
                                dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                         'novedad': 'Domingo y un Festivo manana', 'turno_en_que': domingo_noche,
                                         'num_of_horas': 8}
                                count += 1
                                domingodict[count] = dicts
                            else:
                                dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                         'novedad': domingo, 'turno_en_que': domingo_noche,
                                         'num_of_horas': 2}
                                count += 1
                                domingodict[count] = dicts
                        elif the_date in days_before_holidays:  # Nightshift before a holiday
                            dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                     'novedad': festivo, 'turno_en_que': noche_antes_festivo,
                                     'num_of_horas': 6}
                            count += 1
                            domingodict[count] = dicts
                        elif the_date.weekday() == 5:  # Saturday Night
                            dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                     'novedad': domingo, 'turno_en_que': sabado_noche,
                                     'num_of_horas': 6}
                            count += 1
                            domingodict[count] = dicts

                    elif shift == 'A' or shift == 'B':

                        if the_date in colombianHolidays:
                            dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                     'novedad': festivo, 'turno_en_que': festivo_dia,
                                     'num_of_horas': 8}
                            count += 1
                            domingodict[count] = dicts

                        elif the_date.weekday() == 6:
                            dicts = {'name': name, 'cedula': cedula, 'date': the_date, 'shift': shift,
                                     'novedad': domingo, 'turno_en_que': domingo_dia,
                                     'num_of_horas': 8}
                            count += 1
                            domingodict[count] = dicts

        return domingodict

    def domingo_festivo_report(self,documento='C', nombre='D', novedad='G', turno_enque='H', date='J',
                               num_of_horas='K', starting_row=5, turno='I'):
        domingodict = self.domingo_dict()
        columnC = documento
        columnD = nombre
        columnG = novedad
        columnH = turno_enque
        columnI = turno
        columnJ = date
        columnK = num_of_horas

        row = starting_row  # starting row

        for entry in range(len(domingodict)):
            cellC = columnC + str(entry + row)
            cellD = columnD + str(entry + row)
            cellG = columnG + str(entry + row)
            cellH = columnH + str(entry + row)
            cellI = columnI + str(entry + row)
            cellJ = columnJ + str(entry + row)
            cellK = columnK + str(entry + row)

            self.domingo_festivo.range(cellC).value = domingodict[entry + 1]['cedula']
            self.domingo_festivo.range(cellD).value = domingodict[entry + 1]['name']
            self.domingo_festivo.range(cellG).value = domingodict[entry + 1]['novedad']
            self.domingo_festivo.range(cellH).value = domingodict[entry + 1]['turno_en_que']
            self.domingo_festivo.range(cellI).value = domingodict[entry + 1]['shift']
            self.domingo_festivo.range(cellJ).value = domingodict[entry + 1]['date']
            self.domingo_festivo.range(cellK).value = domingodict[entry + 1]['num_of_horas']

    """----------------------------------------Recargo Nocturno-----------------------------------------------------"""

    def recargo_nocturno_dict(self,row_start=9, col_start='E', name_col='D', cedula_col='C'):
        # initializing count and dictionary
        count = 0
        nocturno_dict = {}

        for i in range(row_start - 1, self.sheet_picked.max_row + 1):  # looping down names
            name = name_col + str(i)
            cedula = cedula_col + str(i)

            name = self.current_schedule.range(name).value
            cedula = self.current_schedule.range(cedula).value

            for x in range(column_index_from_string(col_start), column_index_from_string(self.max_col) + 1,
                           1):  # looping through shifts
                shift = self.sheet_picked.cell(row=i, column=x).value
                date = self.sheet_picked.cell(row=row_start - 2, column=x).value
                if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                    if shift == 'C':
                        if date in colombianHolidays or (date.weekday() == 6):  # nightshift on a sunday or holiday
                            if ((date + datetime.timedelta(days=1)) in colombianHolidays) or (
                                    date.weekday() + 1 == 6):  # the next day is a holiday or sunday as well
                                dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                         'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': ten_to_six,
                                         'num_of_horas': 8}
                                count += 1
                                nocturno_dict[count] = dicts
                            else:  # nightshift on a sunday or holiday with the next day not being a holiday or sunday
                                dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                         'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': ten_to_twelve,
                                         'num_of_horas': 2}
                                count += 1
                                nocturno_dict[count] = dicts

                                dicts = {'name': name, 'cedula': cedula, 'date': (date + datetime.timedelta(days=1)),
                                         'shift': shift,
                                         'tipo_de_recargo': nocturno, 'hora_en_que': twelve_to_six,
                                         'num_of_horas': 6}
                                count += 1
                                nocturno_dict[count] = dicts

                        elif ((date + datetime.timedelta(days=1)) in colombianHolidays) or (
                                date.weekday() + 1 == 6):  # normal nightshift with the next day being a holiday or sunday
                            dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                     'tipo_de_recargo': nocturno, 'hora_en_que': ten_to_twelve,
                                     'num_of_horas': 2}
                            count += 1
                            nocturno_dict[count] = dicts

                            dicts = {'name': name, 'cedula': cedula, 'date': (date + datetime.timedelta(days=1)),
                                     'shift': shift,
                                     'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': twelve_to_six,
                                     'num_of_horas': 6}
                            count += 1
                            nocturno_dict[count] = dicts

                        else:  # normal day night shift, next day not a sunday or a holiday
                            dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                     'tipo_de_recargo': nocturno, 'hora_en_que': ten_to_six,
                                     'num_of_horas': 8}
                            count += 1
                            nocturno_dict[count] = dicts
        return nocturno_dict

    def recargo_nocturno_report(self,documento='C', nombre='D', tipo_de_recargo='G', hora_en_que='I', date='H',
                                num_of_horas='J', starting_row=5):
        nocturno_dict = self.recargo_nocturno_dict()
        columnC = documento
        columnD = nombre
        columnG = tipo_de_recargo
        columnH = date
        columnI = hora_en_que
        columnJ = num_of_horas

        row = starting_row  # starting row

        for entry in range(len(nocturno_dict)):
            cellC = columnC + str(entry + row)
            cellD = columnD + str(entry + row)
            cellG = columnG + str(entry + row)
            cellH = columnH + str(entry + row)
            cellI = columnI + str(entry + row)
            cellJ = columnJ + str(entry + row)

            self.recargo_nocturno.range(cellC).value = nocturno_dict[entry + 1]['cedula']
            self.recargo_nocturno.range(cellD).value = nocturno_dict[entry + 1]['name']
            self.recargo_nocturno.range(cellG).value = nocturno_dict[entry + 1]['tipo_de_recargo']
            self.recargo_nocturno.range(cellI).value = nocturno_dict[entry + 1]['hora_en_que']
            self.recargo_nocturno.range(cellH).value = nocturno_dict[entry + 1]['date']
            self.recargo_nocturno.range(cellJ).value = nocturno_dict[entry + 1]['num_of_horas']

