from tkinter import *
import openpyxl
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from PIL import ImageTk, Image
from normal_personel_report_generator import ExcelData
from tkinter import messagebox

FONT_NAME = "Impact"
MONTH = "Julio"
SHEET = "JULIO"
filename = ""

#Window Label
window = Tk()
window.title("Nominas Creador")
window.config(padx=50,pady=50,bg="white")

#Title
title_label = Label(text="Nominas Excel Creador",font=(FONT_NAME,35,"bold",'underline'),bg="white",)
title_label.grid(row =0, column = 1,columnspan=3,sticky=W)

canvas = Canvas(width=int(1210/4),height=int(562/4),bg="white",highlightthickness=0,highlightcolor="white")

#Getting the logo right
logo = Image.open('mtc.png')
resize_logo = logo.resize((int(1210/4),int(562/4)))
img=ImageTk.PhotoImage(resize_logo)
canvas.create_image((int(1210/4/2),int(562/4/2)),image=img)
canvas.grid(row=0,column=0,sticky=NW)

#Month label and selection
months_label = Label(text="Elija el Mes",bg="white",pady=10,font="bold")
months_label.grid(column=0,row=1)
def month_listbox_used(event):
    # Gets current selection from listbox
    global MONTH
    MONTH = month_listbox.get(month_listbox.curselection())
    months_label.config(text=MONTH)
    print(MONTH)

month_listbox = Listbox(height=12)
months = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
for month in months:
    month_listbox.insert(months.index(month), month)
month_listbox.bind("<<ListboxSelect>>", month_listbox_used)
month_listbox.grid(column=0,row=2,rowspan=4)


def find_schedule_file():
    global filename
    filetypes = (
        ('Excel Files', '*.xlsx'),
        ('All files', '*.*')
    )
    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

    showinfo(
        title='Selected File',
        message=filename
    )
    open_schedule_entry.delete(0,END)
    open_schedule_entry.insert(0,filename)
    sheet_name_label = Label(text="Elija Hoja", bg="white", font=("Colibri", 12, 'bold'))
    sheet_name_label.grid(column=1, columnspan=2, row=3, sticky=S)

    def sheet_listbox_used(event):
        global SHEET
        # Gets current selection from listbox
        sheetname = sheetname_listbox.get(sheetname_listbox.curselection())
        sheet_name_label.config(text="")
        sheet_name_label.config(text=sheetname)
        SHEET = sheetname
        print(sheetname)

    wb = openpyxl.load_workbook(rf"{filename}")
    sheetnames = list(wb.sheetnames)
    sheetname_listbox = Listbox(height=len(sheetnames),width=40)
    for x in sheetnames:
        sheetname_listbox.insert(sheetnames.index(x), x)
    sheetname_listbox.bind("<<ListboxSelect>>", sheet_listbox_used)
    sheetname_listbox.grid(column=1, columnspan=2, row=4,rowspan=5)


# Open schedule file Button and Entry and Label
#Label
open_schedule_label = Label(text="Elija el Archivo de Atendencia",bg="white",font=("Colibri",12,'bold'))
open_schedule_label.grid(column=1,columnspan=2,row=1,sticky=S)
#Entry
open_schedule_entry = Entry(width=50)
open_schedule_entry.grid(column=1,columnspan=2,row=2)
open_schedule_entry.insert(0,"-----------------------Open File--------------------->>>")
#Button
open_schedule_button = Button(text='Encontra el Archivo', command=(find_schedule_file))
open_schedule_button.grid(row=2,column=3,sticky=W)

# Generate reports

def generate_reports():
    excel_workbook = ExcelData(filename,SHEET,MONTH)
    excel_workbook.attendance_creator()
    excel_workbook.attendance_report()
    excel_workbook.domingo_dict()
    excel_workbook.domingo_festivo_report()
    excel_workbook.recargo_nocturno_dict()
    excel_workbook.recargo_nocturno_report()



generate_report = Button(text='GENERAR REPORTE', command=generate_reports,font= 'bold')
generate_report.grid(row=5,column=3,sticky=EW)

def open_template():
    import os
    os.system("start EXCEL.EXE files/template.xlsx")

show_template = Button(text='Abrir Modelo si Necesita Asistencia', command=open_template)
show_template.grid(row=7,column=0,sticky=EW,pady=10)


def open_video():
    import os
    os.system(r"files\video_guide.mp4")

video = Button(text='Video Guía', command=open_video)
video.grid(row=8,column=0,sticky=EW,pady=10)

def open_guide():
    import os
    os.system("start files/Guia.docx")

word_guide = Button(text='Word Guía', command=open_guide)
word_guide.grid(row=9,column=0,sticky=EW,pady=10)







window.mainloop()